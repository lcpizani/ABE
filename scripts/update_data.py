"""
parse_a1_20.py

Parses ISU Extension A1-20 "Estimated Costs of Crop Production in Iowa – 2026"
into normalized records ready for INSERT into PostgreSQL.

Output: list of dicts (or CSV). Each row is one cost line item.

Schema:
    year            INT
    budget_type     TEXT   -- e.g. 'Corn Following Corn'
    sub_budget      TEXT   -- e.g. 'Corn Following Soybeans' (for combined sections like Strip Till)
    section         TEXT   -- 'Preharvest Machinery' | 'Seed Chemical Etc' | 'Harvest Machinery' | 'Labor' | 'Land' | 'Summary'
    cost_item       TEXT   -- e.g. 'Nitrogen, $0.53 per pound'
    cost_type       TEXT   -- 'fixed' | 'variable' | 'total'
    yield_tier      TEXT   -- 'low' | 'mid' | 'high' | None (for machinery table / historical)
    yield_value     NUMERIC -- e.g. 174.0 (bushels or tons per acre)
    yield_unit      TEXT   -- 'bu/acre' | 'ton/acre'
    units_quantity  NUMERIC -- input units (e.g. 28000 kernels, 167 lbs N)
    cost_per_acre   NUMERIC
"""

import re
import csv
import sys
from openpyxl import load_workbook

FILE_PATH = "knowledge/a1-20.xlsx"
OUTPUT_CSV = "crop_costs_2026.csv"
YEAR = 2026

from pathlib import Path
DB_PATH = Path(__file__).parent.parent / "data" / "abe.db"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        return " ".join(v.split())  # collapse whitespace / newlines
    return v


def format_item(label: str) -> str:
    """
    Reformat a cost_item label so that price/unit annotations and footnote
    markers are moved into square brackets.

    Examples:
        "Seed, $3.79 per 1,000 kernels"  →  "Seed [$3.79 per 1,000 kernels]"
        "Nitrogen, $0.53 per pound"       →  "Nitrogen [$0.53 per pound]"
        "Dry (LP gas, $1.54 per gallon)"  →  "Dry [LP gas, $1.54 per gallon]"
        "Preharvest Machinery 1/"         →  "Preharvest Machinery [1/]"
        "Labor"                           →  "Labor"
    """
    if not label:
        return label
    s = " ".join(label.split())

    # 1. Convert parentheticals to brackets — do this BEFORE the comma rule
    #    so "Dry (LP gas, $1.54 per gallon)" becomes "Dry [LP gas, $1.54 per gallon]"
    #    and is not re-processed by the comma rule.
    s = re.sub(r"\s*\(([^)]+)\)", lambda m: f" [{m.group(1).strip()}]", s)

    # 2. Move trailing footnote markers into brackets
    #    "Preharvest Machinery 1/" → "Preharvest Machinery [1/]"
    #    "Annual fertilizer 3/ 0-13-50 …" — footnote is mid-string; bracket it in place
    s = re.sub(r"\s+(\d+/)\s+", lambda m: f" [{m.group(1)}] ", s)  # mid-string
    s = re.sub(r"\s+(\d+/)\s*$", lambda m: f" [{m.group(1)}]", s)  # trailing

    # 3. Move comma-separated price/unit annotations into brackets
    #    "Seed, $3.79 per 1,000 kernels" → "Seed [$3.79 per 1,000 kernels]"
    #    Only fires when the string has no brackets yet (avoids double-bracketing
    #    items already handled by rules 1 or 2) and the tail contains $ or "per".
    if "[" not in s:
        bracket_match = re.match(r"^(.+?),\s+(.+)$", s)
        if bracket_match:
            base = bracket_match.group(1).strip()
            tail = bracket_match.group(2).strip()
            if re.search(r"\$|per ", tail, re.IGNORECASE):
                s = f"{base} [{tail}]"

    return s


def is_section_header(label):
    """True if this looks like a top-level budget section name."""
    keywords = [
        "Corn Following Corn", "Corn Following Soybeans", "Corn Silage Following Corn",
        "Herbicide Tolerant Soybeans Following Corn",
        "Non-Herbicide Tolerant Soybeans Following Corn",
        "Strip Tillage Corn and Soybeans",
        "Low-till Corn and Soybeans",
        "Oats and Hay Production",
        "Annual Production Costs for Established Alfalfa",
        "Maintaining Grass Pastures",
        "Estimated Machinery Costs",
        "Estimated Crop Production Costs in Iowa",
    ]
    return any(label.startswith(k) for k in keywords)


def parse_yield_header(cell_val):
    """'174 bushels per acre' -> (174.0, 'bu/acre')"""
    if not cell_val:
        return None, None
    s = str(cell_val).lower()
    m = re.search(r"([\d.]+)\s*(bushel|bu|ton)", s)
    if not m:
        return None, None
    qty = float(m.group(1))
    unit = "bu/acre" if "bu" in m.group(2) else "ton/acre"
    return qty, unit


def row_dict(rows, idx):
    """Return row as {col_index: value} with Nones stripped."""
    if idx >= len(rows):
        return {}
    return {j: clean(v) for j, v in enumerate(rows[idx]) if v is not None}


# ---------------------------------------------------------------------------
# Main parser for standard 3-tier crop budgets
# Handles: Corn Following Corn, Corn Following Soybeans, Corn Silage Following Corn,
#          HT Soybeans Following Corn, Non-HT Soybeans (each sub-budget separately)
# ---------------------------------------------------------------------------

def find_fixed_var_cols(header_row_dict):
    """
    Given a row like {7:'Fixed', 11:'Variable', 16:'Fixed', ...}
    return list of (fixed_col, variable_col, tier_index) sorted by col position.
    """
    fixed_cols = sorted(c for c, v in header_row_dict.items() if str(v).strip() == "Fixed")
    var_cols   = sorted(c for c, v in header_row_dict.items() if str(v).strip() == "Variable")
    return list(zip(fixed_cols, var_cols))


def parse_standard_budget(rows, start_row, end_row, budget_type, sub_budget=None):
    """
    Parse a standard 3-yield-tier budget block.
    start_row / end_row are 0-based indices into `rows`.
    Returns list of record dicts.
    """
    records = []

    # Row 0 after header: yield values row  (e.g. "174 bushels per acre")
    yield_row = row_dict(rows, start_row)
    # Row 1: Fixed / Variable labels
    fv_row = row_dict(rows, start_row + 1)

    col_pairs = find_fixed_var_cols(fv_row)  # [(fixed_col, var_col), ...]

    # Map each pair to yield info from yield_row
    tiers = []
    for i, (fc, vc) in enumerate(col_pairs):
        # yield value is typically just before the fixed col or at fixed col
        # scan a few cols around fc for the yield label
        yval, yunit = None, None
        for search_col in range(max(0, fc - 3), fc + 4):
            yval, yunit = parse_yield_header(yield_row.get(search_col))
            if yval:
                break
        tier_label = ["low", "mid", "high"][i] if i < 3 else f"tier{i}"
        tiers.append({"label": tier_label, "fc": fc, "vc": vc, "yield": yval, "unit": yunit})

    section = "Preharvest Machinery"
    data_rows = [row_dict(rows, r) for r in range(start_row + 2, end_row)]

    for rd in data_rows:
        if not rd:
            continue
        label = rd.get(0)
        if label is None:
            continue
        label = str(label)

        # Section headers (no numeric values)
        if label in ("Harvest Machinery", "Labor", "Land", "Total fixed, variable"):
            section = label
            continue
        if label.startswith("Seed, Chemical") or label.startswith("Seed, chemical"):
            section = "Seed Chemical Etc"
            continue

        # Skip footnotes and non-data rows
        if label.startswith("1/") or label.startswith("2/") or label.startswith("3/"):
            continue
        if "per acre" in label.lower() and "cash rent" not in label.lower():
            # summary per-acre / per-bushel / per-ton lines
            for t in tiers:
                fv = rd.get(t["fc"])
                vv = rd.get(t["vc"])
                for cost_type, val in [("fixed", fv), ("variable", vv)]:
                    if val is not None and isinstance(val, (int, float)):
                        records.append({
                            "year": YEAR,
                            "budget_type": budget_type,
                            "sub_budget": sub_budget,
                            "section": "Summary",
                            "cost_item": label,
                            "cost_type": cost_type,
                            "yield_tier": t["label"],
                            "yield_value": t["yield"],
                            "yield_unit": t["unit"],
                            "units_quantity": None,
                            "cost_per_acre": val,
                        })
            continue
        if "total cost per" in label.lower():
            for t in tiers:
                val = rd.get(t["fc"])
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        "year": YEAR,
                        "budget_type": budget_type,
                        "sub_budget": sub_budget,
                        "section": "Summary",
                        "cost_item": label,
                        "cost_type": "total",
                        "yield_tier": t["label"],
                        "yield_value": t["yield"],
                        "yield_unit": t["unit"],
                        "units_quantity": None,
                        "cost_per_acre": val,
                    })
            continue

        # Regular cost line
        for t in tiers:
            units_qty = None
            fixed_val = rd.get(t["fc"])
            var_val   = rd.get(t["vc"])

            # units_quantity is the value at fixed_col when it's not a dollar amount
            # (e.g. 28000 kernels, 167 lbs N, 194 lbs P+K) — present only in the first tier
            # Large values (>500) are clearly units; fertilizer rows may have small lbs values
            # so we also detect by label keywords.
            _is_units_row = (
                label and (
                    "per pound" in label.lower()
                    or "per 1,000 kernels" in label.lower()
                    or label.lower().startswith("nitrogen (n)")
                    or label.lower().startswith("phosphate (p")
                    or label.lower().startswith("potash (k")
                    or label.lower().startswith("lime")
                )
            )
            if isinstance(fixed_val, (int, float)) and (fixed_val > 500 or _is_units_row):
                units_qty = fixed_val
                fixed_val = None  # no separate fixed cost listed for units rows

            for cost_type, val in [("fixed", fixed_val), ("variable", var_val)]:
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        "year": YEAR,
                        "budget_type": budget_type,
                        "sub_budget": sub_budget,
                        "section": section,
                        "cost_item": label,
                        "cost_type": cost_type,
                        "yield_tier": t["label"],
                        "yield_value": t["yield"],
                        "yield_unit": t["unit"],
                        "units_quantity": units_qty,
                        "cost_per_acre": val,
                    })

    return records


# ---------------------------------------------------------------------------
# Parser for paired side-by-side budgets (Strip Till, Low-Till)
# These have two crops next to each other on the same rows.
# ---------------------------------------------------------------------------

def parse_paired_budget(rows, start_row, end_row, budget_type, left_sub, right_sub):
    """
    Rows like Strip Tillage or Low-Till have left (corn) and right (soybean) budgets
    side by side. We parse each half independently.
    """
    records = []
    for sub, side in [(left_sub, "left"), (right_sub, "right")]:
        records += _parse_paired_side(rows, start_row, end_row, budget_type, sub, side)
    return records


def _parse_paired_side(rows, start_row, end_row, budget_type, sub_budget, side):
    records = []

    # Find yield row (row after the sub-budget label rows)
    yield_row = row_dict(rows, start_row + 1)
    fv_row    = row_dict(rows, start_row + 2)

    all_fixed = sorted(c for c, v in fv_row.items() if str(v).strip() == "Fixed")
    all_var   = sorted(c for c, v in fv_row.items() if str(v).strip() == "Variable")

    if not all_fixed:
        return records

    # Split halves at the midpoint column
    mid = (max(all_fixed) + min(all_fixed)) // 2
    if side == "left":
        fixed_cols = [c for c in all_fixed if c <= mid]
        var_cols   = [c for c in all_var   if c <= mid]
    else:
        fixed_cols = [c for c in all_fixed if c > mid]
        var_cols   = [c for c in all_var   if c > mid]

    col_pairs = list(zip(fixed_cols, var_cols))
    tiers = []
    for i, (fc, vc) in enumerate(col_pairs):
        yval, yunit = None, None
        for sc in range(max(0, fc - 4), fc + 5):
            yval, yunit = parse_yield_header(yield_row.get(sc))
            if yval:
                break
        tier_label = ["low", "mid", "high"][i] if i < 3 else f"tier{i}"
        tiers.append({"label": tier_label, "fc": fc, "vc": vc, "yield": yval, "unit": yunit})

    section = "Preharvest Machinery"
    for r in range(start_row + 3, end_row):
        rd = row_dict(rows, r)
        if not rd:
            continue

        # For paired budgets, the right side replicates label in col ~17; use col 0 for left
        label = rd.get(0)
        if label is None:
            continue
        label = str(label)

        if label.startswith("1/") or label.startswith("2/"):
            continue
        if label in ("Harvest Machinery", "Labor", "Land", "Total fixed, variable"):
            section = label
            continue
        if label.startswith("Seed, Chemical") or label.startswith("Seed, chemical"):
            section = "Seed Chemical Etc"
            continue

        for t in tiers:
            units_qty = None
            fixed_val = rd.get(t["fc"])
            var_val   = rd.get(t["vc"])

            if isinstance(fixed_val, (int, float)) and fixed_val > 500:
                units_qty = fixed_val
                fixed_val = None

            for cost_type, val in [("fixed", fixed_val), ("variable", var_val)]:
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        "year": YEAR,
                        "budget_type": budget_type,
                        "sub_budget": sub_budget,
                        "section": section,
                        "cost_item": label,
                        "cost_type": cost_type,
                        "yield_tier": t["label"],
                        "yield_value": t["yield"],
                        "yield_unit": t["unit"],
                        "units_quantity": units_qty,
                        "cost_per_acre": val,
                    })
    return records


# ---------------------------------------------------------------------------
# Parser for Machinery Costs table (rows 314-357)
# ---------------------------------------------------------------------------

def parse_machinery_table(rows, start_row, end_row):
    records = []
    # Header is at start_row+1 (row 316): Operation | Hours | Fixed | Variable
    for r in range(start_row + 2, end_row):
        rd = row_dict(rows, r)
        if not rd:
            continue
        label = rd.get(0)
        if not label or not isinstance(label, str):
            continue
        label = label.strip()
        if not label or label.startswith("The following"):
            continue

        hours = rd.get(4)
        fixed_val = rd.get(17)
        var_val   = rd.get(28)
        unit_col  = rd.get(21) or rd.get(33) or "/acre"

        for cost_type, val in [("fixed", fixed_val), ("variable", var_val)]:
            if val is not None and isinstance(val, (int, float)):
                records.append({
                    "year": YEAR,
                    "budget_type": "Estimated Machinery Costs",
                    "sub_budget": None,
                    "section": "Machinery",
                    "cost_item": label,
                    "cost_type": cost_type,
                    "yield_tier": None,
                    "yield_value": None,
                    "yield_unit": str(unit_col).strip() if unit_col else None,
                    "units_quantity": hours,
                    "cost_per_acre": val,
                })
    return records


# ---------------------------------------------------------------------------
# Parser for historical trend table (rows 358-392)
# ---------------------------------------------------------------------------

def parse_historical_table(rows, start_row, end_row):
    records = []
    # Row 359 has the years across columns
    year_row = row_dict(rows, start_row + 1)
    years = {col: int(str(v).replace("1/", "").replace("2/", "").strip())
             for col, v in year_row.items()
             if isinstance(v, (int, float)) or (isinstance(v, str) and re.match(r"\d{4}", v.strip()[:4]))}

    budget = None
    for r in range(start_row + 2, end_row):
        rd = row_dict(rows, r)
        if not rd:
            continue
        label = rd.get(0)
        if not label:
            continue
        label = str(label).strip()
        if label.startswith("1/") or label.startswith("2/") or label.startswith("3/"):
            continue
        if is_section_header(label) or label in (
            "Corn Following Corn", "Corn Following Soybeans",
            "Soybeans Following Corn 3/", "Alfalfa Hay, annual production, 6 ton per acre, large round bales"
        ):
            budget = label
            continue
        for col, yr in years.items():
            val = rd.get(col)
            if val is not None and isinstance(val, (int, float)):
                records.append({
                    "year": yr,
                    "budget_type": budget,
                    "sub_budget": None,
                    "section": "Historical Summary",
                    "cost_item": label,
                    "cost_type": "total",
                    "yield_tier": None,
                    "yield_value": None,
                    "yield_unit": None,
                    "units_quantity": None,
                    "cost_per_acre": val,
                })
    return records


# ---------------------------------------------------------------------------
# Alfalfa / Hay Establishment (rows 230-267)
# ---------------------------------------------------------------------------

def parse_alfalfa_establishment(rows, start_row, end_row):
    records = []
    fv_row = row_dict(rows, start_row + 1)  # row 231: sub-budget labels
    fv_row2 = row_dict(rows, start_row + 2)  # row 232: Fixed / Variable

    # Two sub-budgets side by side
    # Left: Alfalfa-Grass Seeded with Oat (col 8/17), Right: Alfalfa Seeded with Herbicide (col 20/27)
    configs = [
        ("Alfalfa-Grass Seeded with Oat Companion Crop", 8, 17),
        ("Alfalfa Seeded with Herbicide", 20, 27),
    ]
    section = "Establishment"
    for r in range(start_row + 3, end_row):
        rd = row_dict(rows, r)
        if not rd:
            continue
        label = rd.get(0)
        if not label:
            continue
        label = str(label).strip()
        if label.startswith("1/") or label.startswith("2/") or label.startswith("3/") or label.startswith("4/"):
            continue
        if label in ("Annual Costs", "Harvest Machinery", "Seed 3/", "Fertilizer"):
            section = label
            continue

        for sub, fc, vc in configs:
            fv = rd.get(fc)
            vv = rd.get(vc)
            for cost_type, val in [("fixed", fv), ("variable", vv)]:
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        "year": YEAR,
                        "budget_type": "Oats and Hay Production - Seeding Year",
                        "sub_budget": sub,
                        "section": section,
                        "cost_item": label,
                        "cost_type": cost_type,
                        "yield_tier": None,
                        "yield_value": None,
                        "yield_unit": None,
                        "units_quantity": None,
                        "cost_per_acre": val,
                    })
    return records


# ---------------------------------------------------------------------------
# Alfalfa Annual Production (rows 268-290)
# ---------------------------------------------------------------------------

def parse_alfalfa_annual(rows, start_row, end_row):
    records = []
    yield_row = row_dict(rows, start_row + 2)   # row 270: 4 tons / 6 tons
    fv_row    = row_dict(rows, start_row + 3)   # row 271: Fixed / Variable

    col_pairs = find_fixed_var_cols(fv_row)
    tiers = []
    for i, (fc, vc) in enumerate(col_pairs):
        yval, yunit = None, None
        for sc in range(max(0, fc - 3), fc + 4):
            yval, yunit = parse_yield_header(yield_row.get(sc))
            if yval:
                break
        tiers.append({"label": f"{int(yval) if yval else i}ton", "fc": fc, "vc": vc, "yield": yval, "unit": yunit or "ton/acre"})

    section = "Annual Production"
    for r in range(start_row + 4, end_row):
        rd = row_dict(rows, r)
        if not rd:
            continue
        label = rd.get(0)
        if not label:
            continue
        label = str(label).strip()
        if label.startswith("1/") or label.startswith("2/") or label.startswith("3/") or label.startswith("4/"):
            continue
        if label.startswith("Harvesting Costs:"):
            section = label
            continue

        for t in tiers:
            fv = rd.get(t["fc"])
            vv = rd.get(t["vc"])
            for cost_type, val in [("fixed", fv), ("variable", vv)]:
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        "year": YEAR,
                        "budget_type": "Annual Production Costs for Established Alfalfa or Alfalfa-Grass Hay",
                        "sub_budget": None,
                        "section": section,
                        "cost_item": label,
                        "cost_type": cost_type,
                        "yield_tier": t["label"],
                        "yield_value": t["yield"],
                        "yield_unit": t["unit"],
                        "units_quantity": None,
                        "cost_per_acre": val,
                    })
    return records


# ---------------------------------------------------------------------------
# Grass Pasture (rows 291-313)
# ---------------------------------------------------------------------------

def parse_pasture(rows, start_row, end_row):
    records = []
    sub_row = row_dict(rows, start_row + 1)   # row 292: sub-budget names
    fv_row  = row_dict(rows, start_row + 2)   # row 293: Fixed / Variable

    configs = [
        ("Improved Grass", 5, 12),
        ("Improved Grass-Legume", 19, 24),
    ]
    section = "Pasture"
    for r in range(start_row + 3, end_row):
        rd = row_dict(rows, r)
        if not rd:
            continue
        label = rd.get(0)
        if not label:
            continue
        label = str(label).strip()
        if label.startswith("1/") or label.startswith("2/") or label.startswith("3/"):
            continue
        if label in ("Machinery Costs", "Fertilizer and Herbicide 1/", "Labor", "Land"):
            section = label
            continue

        for sub, fc, vc in configs:
            fv = rd.get(fc)
            vv = rd.get(vc)
            for cost_type, val in [("fixed", fv), ("variable", vv)]:
                if val is not None and isinstance(val, (int, float)):
                    records.append({
                        "year": YEAR,
                        "budget_type": "Maintaining Grass Pastures",
                        "sub_budget": sub,
                        "section": section,
                        "cost_item": label,
                        "cost_type": cost_type,
                        "yield_tier": None,
                        "yield_value": None,
                        "yield_unit": None,
                        "units_quantity": None,
                        "cost_per_acre": val,
                    })
    return records


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def parse_all(file_path):
    wb = load_workbook(file_path, read_only=True)
    ws = wb["Table 1"]
    rows = list(ws.iter_rows(values_only=True))

    all_records = []

    # Standard 3-tier budgets (0-indexed row numbers = xlsx row - 1)
    # Row index points to the yield-header row (one below the budget title row).
    # Corn Following Corn: title row 5, yield header row 6 -> 0-indexed 6, end 37
    all_records += parse_standard_budget(rows, 6, 37, "Corn Following Corn")

    # Corn Following Soybeans: title row 38, yield header row 39 -> 0-indexed 39, end 69
    all_records += parse_standard_budget(rows, 39, 69, "Corn Following Soybeans")

    # Corn Silage Following Corn: title row 70, yield header row 71 -> 0-indexed 71, end 100
    all_records += parse_standard_budget(rows, 71, 100, "Corn Silage Following Corn")

    # HT Soybeans Following Corn: title row 101, yield header row 102 -> 0-indexed 102, end 130
    all_records += parse_standard_budget(rows, 102, 130, "Herbicide Tolerant Soybeans Following Corn")

    # Strip Tillage (paired): rows 132-163 -> 131-162
    all_records += parse_paired_budget(
        rows, 132, 163,
        "Strip Tillage",
        "Corn Following Soybeans",
        "Herbicide Tolerant Soybeans Following Corn",
    )

    # Non-HT Soybeans Following Corn: rows 165-196 -> 164-195
    # This has two sub-budgets side by side: Soybeans Following Corn + Drilled Soybeans
    all_records += parse_paired_budget(
        rows, 165, 195,
        "Non-Herbicide Tolerant Soybeans Following Corn",
        "Soybeans Following Corn",
        "Drilled Soybeans Following Corn",
    )

    # Low-Till (paired): rows 197-229 -> 196-228
    all_records += parse_paired_budget(
        rows, 197, 228,
        "Low-Till",
        "Corn Following Soybeans",
        "Herbicide Tolerant Drilled Soybeans Following Corn",
    )

    # Alfalfa Establishment: rows 230-267 -> 229-266
    all_records += parse_alfalfa_establishment(rows, 229, 266)

    # Alfalfa Annual Production: rows 268-290 -> 267-289
    all_records += parse_alfalfa_annual(rows, 267, 289)

    # Grass Pasture: rows 291-313 -> 290-312
    all_records += parse_pasture(rows, 290, 312)

    # Machinery Costs table: rows 314-357 -> 313-356
    all_records += parse_machinery_table(rows, 313, 356)

    # Historical trend table: rows 358-392 -> 357-391
    all_records += parse_historical_table(rows, 357, 391)

    # Normalize cost_item labels across all records
    for r in all_records:
        r["cost_item"] = format_item(r["cost_item"])

    return all_records


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

FIELDS = [
    "year", "budget_type", "sub_budget", "section",
    "cost_item", "cost_type", "yield_tier", "yield_value",
    "yield_unit", "units_quantity", "cost_per_acre",
]


def to_csv(records, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(records)
    print(f"Wrote {len(records)} rows to {path}")


def to_sqlite(records, db_path=None):
    """Insert all records into data/abe.db as crop_production_costs table."""
    import sqlite3
    path = db_path or DB_PATH
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(path) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS crop_production_costs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                year            INTEGER,
                budget_type     TEXT,
                sub_budget      TEXT,
                section         TEXT,
                cost_item       TEXT,
                cost_type       TEXT,
                yield_tier      TEXT,
                yield_value     REAL,
                yield_unit      TEXT,
                units_quantity  REAL,
                cost_per_acre   REAL
            )
        """)
        conn.execute("DELETE FROM crop_production_costs")
        conn.executemany(
            """
            INSERT INTO crop_production_costs
                (year, budget_type, sub_budget, section, cost_item, cost_type,
                 yield_tier, yield_value, yield_unit, units_quantity, cost_per_acre)
            VALUES
                (:year, :budget_type, :sub_budget, :section, :cost_item, :cost_type,
                 :yield_tier, :yield_value, :yield_unit, :units_quantity, :cost_per_acre)
            """,
            records,
        )
        count = conn.execute("SELECT COUNT(*) FROM crop_production_costs").fetchone()[0]
    print(f"Seeded {count} rows into crop_production_costs in {path}")


def to_postgres_insert(records):
    """Print INSERT statements (or adapt to psycopg2 executemany)."""
    print("-- Paste into psql or run via psycopg2\n")
    print("""CREATE TABLE IF NOT EXISTS crop_production_costs (
    id              SERIAL PRIMARY KEY,
    year            INT,
    budget_type     TEXT,
    sub_budget      TEXT,
    section         TEXT,
    cost_item       TEXT,
    cost_type       TEXT,
    yield_tier      TEXT,
    yield_value     NUMERIC,
    yield_unit      TEXT,
    units_quantity  NUMERIC,
    cost_per_acre   NUMERIC
);\n""")

    vals = []
    for r in records:
        def q(v):
            if v is None:
                return "NULL"
            if isinstance(v, (int, float)):
                return str(v)
            return "'" + str(v).replace("'", "''") + "'"

        vals.append(
            f"({q(r['year'])},{q(r['budget_type'])},{q(r['sub_budget'])},{q(r['section'])},"
            f"{q(r['cost_item'])},{q(r['cost_type'])},{q(r['yield_tier'])},{q(r['yield_value'])},"
            f"{q(r['yield_unit'])},{q(r['units_quantity'])},{q(r['cost_per_acre'])})"
        )

    chunk = 500
    for i in range(0, len(vals), chunk):
        batch = vals[i:i+chunk]
        print("INSERT INTO crop_production_costs "
              "(year,budget_type,sub_budget,section,cost_item,cost_type,"
              "yield_tier,yield_value,yield_unit,units_quantity,cost_per_acre) VALUES")
        print(",\n".join(batch) + ";")


if __name__ == "__main__":
    import os
    src = FILE_PATH
    if not os.path.exists(src):
        src = "knowledge/a1-20.xlsx"

    records = parse_all(src)

    mode = sys.argv[1] if len(sys.argv) > 1 else "db"

    if mode == "sql":
        to_postgres_insert(records)
    elif mode == "csv":
        to_csv(records, OUTPUT_CSV)
        print("\nSample rows:")
        for r in records[:5]:
            print(r)
    else:
        to_sqlite(records)